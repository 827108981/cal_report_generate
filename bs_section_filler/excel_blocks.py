from __future__ import annotations

import math
import re
import statistics
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

from .utils import (
    canonical_section,
    compact_texts,
    format_number_by_excel_format,
    is_blank_value,
    looks_like_section,
    normalize_text,
)


@dataclass
class SourceCell:
    value: Any
    row: int
    col: int
    is_formula: bool = False
    number_format: str = "General"

    @property
    def text(self) -> str:
        return format_number_by_excel_format(self.value, self.number_format)

    @property
    def norm(self) -> str:
        return normalize_text(self.text)


@dataclass
class ExcelBlock:
    module: str
    sheet: str
    section: str
    title: str
    object_name: str
    start_row: int
    end_row: int
    rows: list[list[SourceCell]]
    header_norms: set[str]

    @property
    def key_text(self) -> str:
        return " ".join(x for x in [self.section, self.title, self.object_name] if x)


def _cell_is_formula(cell: Cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def _to_number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        if text.endswith("%"):
            return float(text[:-1]) / 100
        return float(text)
    except ValueError:
        return None


def _split_args(text: str) -> list[str]:
    args: list[str] = []
    start = 0
    depth = 0
    in_quote = False
    for i, ch in enumerate(text):
        if ch == '"':
            in_quote = not in_quote
        elif not in_quote:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            elif ch == "," and depth == 0:
                args.append(text[start:i].strip())
                start = i + 1
    args.append(text[start:].strip())
    return args


def _split_formula_concat(text: str) -> list[str]:
    parts: list[str] = []
    start = 0
    in_quote = False
    for i, ch in enumerate(text):
        if ch == '"':
            in_quote = not in_quote
        elif ch == "&" and not in_quote:
            parts.append(text[start:i].strip())
            start = i + 1
    parts.append(text[start:].strip())
    return parts


def _build_formula_evaluator(ws_val, ws_formula):
    """Evaluate the subset of Excel formulas used by the calibration workbooks.

    openpyxl does not calculate formulas. The workbook committed in the repository
    contains several stale/error cached values, so the report generator needs a
    deterministic evaluator. The evaluator intentionally supports only safe,
    known spreadsheet operations and never calls eval().
    """

    cache: dict[tuple[int, int], Any] = {}
    visiting: set[tuple[int, int]] = set()

    def strip_outer_parens(expr: str) -> str:
        e = expr.strip()
        changed = True
        while changed and e.startswith("(") and e.endswith(")"):
            changed = False
            depth = 0
            in_quote = False
            for i, ch in enumerate(e):
                if ch == '"':
                    in_quote = not in_quote
                elif not in_quote:
                    if ch == "(":
                        depth += 1
                    elif ch == ")":
                        depth -= 1
                    if depth == 0 and i != len(e) - 1:
                        return e
            if depth == 0:
                e = e[1:-1].strip()
                changed = True
        return e

    def cell_value(ref: str) -> Any:
        row, col = coordinate_to_tuple(ref.replace("$", ""))
        return evaluate(row, col)

    def cell_number(ref: str) -> float | None:
        return _to_number(cell_value(ref))

    def range_values(ref: str) -> list[Any]:
        min_col, min_row, max_col, max_row = range_boundaries(ref.replace("$", ""))
        return [evaluate(rr, cc) for rr in range(min_row, max_row + 1) for cc in range(min_col, max_col + 1)]

    def range_numbers(ref: str) -> list[float]:
        numbers: list[float] = []
        for value in range_values(ref):
            number = _to_number(value)
            if number is not None:
                numbers.append(number)
        return numbers

    def average(ref: str) -> float | None:
        numbers = range_numbers(ref)
        return sum(numbers) / len(numbers) if numbers else None

    def stdev(ref: str) -> float | None:
        numbers = range_numbers(ref)
        return statistics.stdev(numbers) if len(numbers) >= 2 else None

    def slope(y_ref: str, x_ref: str) -> float | None:
        ys = range_numbers(y_ref)
        xs = range_numbers(x_ref)
        pairs = list(zip(xs, ys))
        if len(pairs) < 2:
            return None
        mean_x = sum(x for x, _ in pairs) / len(pairs)
        mean_y = sum(y for _, y in pairs) / len(pairs)
        denominator = sum((x - mean_x) ** 2 for x, _ in pairs)
        if denominator == 0:
            return None
        return sum((x - mean_x) * (y - mean_y) for x, y in pairs) / denominator

    def intercept(y_ref: str, x_ref: str) -> float | None:
        m = slope(y_ref, x_ref)
        xs = range_numbers(x_ref)
        ys = range_numbers(y_ref)
        if m is None or not xs or not ys:
            return None
        return sum(ys) / len(ys) - m * (sum(xs) / len(xs))

    def correl(xs: list[float], ys: list[float]) -> float | None:
        pairs = list(zip(xs, ys))
        if len(pairs) < 2:
            return None
        mean_x = sum(x for x, _ in pairs) / len(pairs)
        mean_y = sum(y for _, y in pairs) / len(pairs)
        sx = math.sqrt(sum((x - mean_x) ** 2 for x, _ in pairs))
        sy = math.sqrt(sum((y - mean_y) ** 2 for _, y in pairs))
        if sx == 0 or sy == 0:
            return None
        return sum((x - mean_x) * (y - mean_y) for x, y in pairs) / (sx * sy)

    def split_top_level_operator(expr: str, operators: set[str]) -> tuple[str, str, str] | None:
        depth = 0
        in_quote = False
        for i in range(len(expr) - 1, -1, -1):
            ch = expr[i]
            if ch == '"':
                in_quote = not in_quote
                continue
            if in_quote:
                continue
            if ch == ")":
                depth += 1
                continue
            if ch == "(":
                depth -= 1
                continue
            if depth != 0 or ch not in operators:
                continue
            previous = expr[i - 1] if i > 0 else ""
            if ch in "+-" and (i == 0 or previous in "+-*/(<>=,"):
                continue
            return expr[:i], ch, expr[i + 1 :]
        return None

    def split_top_level_comparison(expr: str) -> tuple[str, str, str] | None:
        depth = 0
        in_quote = False
        i = 0
        while i < len(expr):
            ch = expr[i]
            if ch == '"':
                in_quote = not in_quote
                i += 1
                continue
            if in_quote:
                i += 1
                continue
            if ch == "(":
                depth += 1
                i += 1
                continue
            if ch == ")":
                depth -= 1
                i += 1
                continue
            if depth == 0:
                pair = expr[i : i + 2]
                if pair in {"<=", ">=", "<>"}:
                    return expr[:i], pair, expr[i + 2 :]
                if ch in {"=", "<", ">"}:
                    return expr[:i], ch, expr[i + 1 :]
            i += 1
        return None

    def countif(range_ref: str, criterion_expr: str) -> int:
        criterion = scalar_expr(criterion_expr)
        if criterion is None:
            criterion = criterion_expr.strip().strip('"')
        criterion_text = str(criterion).strip()
        operator = "="
        right_text = criterion_text
        for prefix in ("<=", ">=", "<>", "<", ">", "="):
            if criterion_text.startswith(prefix):
                operator = prefix
                right_text = criterion_text[len(prefix) :].strip()
                break
        if right_text == "":
            return sum(1 for value in range_values(range_ref) if is_blank_value(value))
        right_number = _to_number(right_text)
        total = 0
        for value in range_values(range_ref):
            if operator == "<>":
                if is_blank_value(value):
                    total += 1
                elif right_number is not None:
                    left_number = _to_number(value)
                    if left_number is None or left_number != right_number:
                        total += 1
                elif str(value).strip() != right_text:
                    total += 1
                continue
            if right_number is not None:
                left_number = _to_number(value)
                if left_number is None:
                    continue
                ok = (
                    (operator == "=" and left_number == right_number)
                    or (operator == "<" and left_number < right_number)
                    or (operator == ">" and left_number > right_number)
                    or (operator == "<=" and left_number <= right_number)
                    or (operator == ">=" and left_number >= right_number)
                )
            else:
                left_text = "" if is_blank_value(value) else str(value).strip()
                ok = operator == "=" and left_text == right_text
            if ok:
                total += 1
        return total

    def scalar_expr(expr: str) -> Any:
        e = strip_outer_parens(expr)
        if e.startswith('"') and e.endswith('"'):
            return e[1:-1]
        if re.fullmatch(r"-?\d+(?:\.\d+)?%?", e):
            return float(e[:-1]) / 100 if e.endswith("%") else float(e)

        split = split_top_level_operator(e, {"+", "-"})
        if split:
            left = _to_number(scalar_expr(split[0]))
            right = _to_number(scalar_expr(split[2]))
            if left is None or right is None:
                return None
            return left + right if split[1] == "+" else left - right

        split = split_top_level_operator(e, {"*", "/"})
        if split:
            left = _to_number(scalar_expr(split[0]))
            right = _to_number(scalar_expr(split[2]))
            if left is None or right is None:
                return None
            if split[1] == "/":
                return left / right if right != 0 else None
            return left * right

        if re.fullmatch(r"\$?[A-Z]{1,3}\$?\d+", e, flags=re.I):
            return cell_value(e)

        match = re.fullmatch(r"ABS\((.+)\)", e, flags=re.I)
        if match:
            number = _to_number(scalar_expr(match.group(1)))
            return abs(number) if number is not None else None
        match = re.fullmatch(r"SUM\(([^)]+)\)", e, flags=re.I)
        if match:
            # Excel SUM ignores blanks. Returning 0 for an all-blank range matches Excel.
            return sum(range_numbers(match.group(1)))
        match = re.fullmatch(r"AVERAGE\(([^)]+)\)", e, flags=re.I)
        if match:
            return average(match.group(1))
        match = re.fullmatch(r"STDEV(?:\.S)?\(([^)]+)\)", e, flags=re.I)
        if match:
            return stdev(match.group(1))
        match = re.fullmatch(r"MAX\(([^)]+)\)", e, flags=re.I)
        if match:
            numbers = range_numbers(match.group(1))
            return max(numbers) if numbers else None
        match = re.fullmatch(r"MIN\(([^)]+)\)", e, flags=re.I)
        if match:
            numbers = range_numbers(match.group(1))
            return min(numbers) if numbers else None
        match = re.fullmatch(r"COUNTIF\(([^,]+),\s*(.+)\)", e, flags=re.I)
        if match:
            return countif(match.group(1), match.group(2))
        return None

    def compare_expr(condition: str) -> bool:
        comparison = split_top_level_comparison(strip_outer_parens(condition))
        if not comparison:
            return bool(scalar_expr(condition))
        left = scalar_expr(comparison[0])
        right = scalar_expr(comparison[2])
        operator = comparison[1]
        if operator in {"=", "<>"} and (_to_number(left) is None or _to_number(right) is None):
            left_text = "" if is_blank_value(left) else str(left).strip()
            right_text = "" if is_blank_value(right) else str(right).strip()
            return (left_text == right_text) if operator == "=" else (left_text != right_text)
        left_number = _to_number(left)
        right_number = _to_number(right)
        if left_number is None or right_number is None:
            return False
        if operator == "<=":
            return left_number <= right_number
        if operator == ">=":
            return left_number >= right_number
        if operator == "<":
            return left_number < right_number
        if operator == ">":
            return left_number > right_number
        if operator == "<>":
            return left_number != right_number
        return left_number == right_number

    def condition_expr(expr: str) -> bool:
        e = strip_outer_parens(expr)
        upper = e.upper()
        if upper.startswith("AND(") and e.endswith(")"):
            return all(condition_expr(arg) for arg in _split_args(e[4:-1]))
        if upper.startswith("OR(") and e.endswith(")"):
            return any(condition_expr(arg) for arg in _split_args(e[3:-1]))
        return compare_expr(e)

    def evaluate_formula(formula: str) -> Any:
        f = formula.strip()
        if f.startswith("="):
            f = f[1:].strip()

        for name, func in (
            ("SUM", lambda ref: sum(range_numbers(ref))),
            ("AVERAGE", average),
            ("STDEV", stdev),
            ("STDEV.S", stdev),
        ):
            match = re.fullmatch(rf"{re.escape(name)}\(([^)]+)\)", f, flags=re.I)
            if match:
                return func(match.group(1))
        match = re.fullmatch(r"MIN\(([^)]+)\)", f, flags=re.I)
        if match:
            numbers = range_numbers(match.group(1))
            return min(numbers) if numbers else None
        match = re.fullmatch(r"MAX\(([^)]+)\)", f, flags=re.I)
        if match:
            numbers = range_numbers(match.group(1))
            return max(numbers) if numbers else None
        match = re.fullmatch(r"COUNTIF\(([^,]+),\s*(.+)\)", f, flags=re.I)
        if match:
            return countif(match.group(1), match.group(2))
        match = re.fullmatch(r"POWER\(([^,]+),\s*([^)]+)\)", f, flags=re.I)
        if match:
            base = _to_number(scalar_expr(match.group(1)))
            exponent = _to_number(scalar_expr(match.group(2)))
            return base**exponent if base is not None and exponent is not None else None
        match = re.fullmatch(r"SLOPE\(([^,]+),\s*([^)]+)\)", f, flags=re.I)
        if match:
            return slope(match.group(1), match.group(2))
        match = re.fullmatch(r"INTERCEPT\(([^,]+),\s*([^)]+)\)", f, flags=re.I)
        if match:
            return intercept(match.group(1), match.group(2))

        match = re.fullmatch(
            r"CORREL\(([^,]+),\s*\(([^)]+)\)\s*\*\s*(\$?[A-Z]{1,3}\$?\d+)\s*\+\s*(\$?[A-Z]{1,3}\$?\d+)\)",
            f,
            flags=re.I,
        )
        if match:
            ys = range_numbers(match.group(1))
            xs = range_numbers(match.group(2))
            a = cell_number(match.group(3))
            b = cell_number(match.group(4))
            if a is None or b is None:
                return None
            return correl(ys, [x * a + b for x in xs])

        if f.upper().startswith("IF(") and f.endswith(")"):
            args = _split_args(f[3:-1])
            if len(args) >= 3:
                # Do not turn an empty/unexecuted template block into a false Fail.
                # BS-5000 ISE module 2 contains formulas but no measurements; Excel
                # caches #DIV/0! there. If any directly referenced condition cell is
                # unavailable, keep the result blank instead of synthesizing Fail.
                refs = re.findall(r"\$?[A-Z]{1,3}\$?\d+", args[0], flags=re.I)
                if refs and any(is_blank_value(cell_value(ref)) for ref in refs):
                    return None
                return scalar_expr(args[1]) if condition_expr(args[0]) else scalar_expr(args[2])

        value = scalar_expr(f)
        if value is not None:
            return value
        if "&" in f:
            parts = _split_formula_concat(f)
            return "".join(format_number_by_excel_format(scalar_expr(part)) for part in parts)
        return None

    def evaluate(row: int, col: int) -> Any:
        key = (row, col)
        if key in cache:
            return cache[key]
        raw = ws_val.cell(row, col).value
        formula = ws_formula.cell(row, col).value
        if not (isinstance(formula, str) and formula.startswith("=")):
            cache[key] = raw
            return raw
        if key in visiting:
            return raw if not is_blank_value(raw) else None
        visiting.add(key)
        try:
            calculated = evaluate_formula(formula)
            if calculated is not None:
                cache[key] = calculated
                return calculated
            cache[key] = raw
            return raw
        finally:
            visiting.discard(key)

    return evaluate


def _compress_row(ws_val, ws_formula, row: int, max_col: int, evaluator=None) -> list[SourceCell]:
    result: list[SourceCell] = []
    for col in range(1, max_col + 1):
        value = ws_val.cell(row, col).value
        formula = ws_formula.cell(row, col).value
        if evaluator is not None and isinstance(formula, str) and formula.startswith("="):
            calculated = evaluator(row, col)
            if calculated is not None:
                value = calculated
        if is_blank_value(value):
            continue
        result.append(
            SourceCell(
                value=value,
                row=row,
                col=col,
                is_formula=isinstance(formula, str) and formula.startswith("="),
                number_format=ws_formula.cell(row, col).number_format or ws_val.cell(row, col).number_format or "General",
            )
        )
    return result


def _row_text(row: list[SourceCell]) -> str:
    return " ".join(cell.text for cell in row if cell.text)


def _module_from_text(text: str) -> str | None:
    normalized = normalize_text(text)
    if "模块二" in normalized or "M2模块" in normalized or "模块2" in normalized or "ISE模块二" in normalized or "ISE模块2" in normalized:
        return "M2"
    if "模块一" in normalized or "M1模块" in normalized or "模块1" in normalized or "ISE模块一" in normalized or "ISE模块1" in normalized:
        return "M1"
    return None


def _is_module_marker(row: list[SourceCell]) -> bool:
    texts = compact_texts(cell.text for cell in row)
    return len(texts) == 1 and _module_from_text(texts[0]) is not None


def _is_anchor(row: list[SourceCell]) -> bool:
    if not row or _is_module_marker(row):
        return False
    texts = [cell.text for cell in row]
    first = texts[0]
    joined = _row_text(row)
    normalized = normalize_text(joined)
    if first.startswith("注") or any(value in first for value in ["测试次数", "测试项目"]):
        return False
    if "试剂名称" in joined:
        return True
    if any(
        keyword in joined
        for keyword in [
            "测试结果",
            "吸光度线性值",
            "吸光度准确性",
            "暗电流测试",
            "真空及压力检测",
            "工作环境检测",
            "仪器维护保养",
            "仪器基本状态检查",
            "主机位置校准",
            "SDM位置校准",
            "反应盘温度检测",
            "样本携带污染率检测",
            "电解质准确度",
            "电解质精密度",
            "电解质线性检测",
            "线性范围验证",
            "电解质稳定性",
            "电解质携带污染率",
            "项目交叉污染",
            "项目测试时间",
            "杂散光",
            "亚硝酸钠",
        ]
    ):
        return True
    if len(texts) == 1:
        if any(
            value in first
            for value in [
                "批号",
                "重复次数",
                "测试数据",
                "次数",
                "采光周期",
                "要求",
                "结论",
                "指标",
                "统计",
                "R²",
                "R：",
                "a：",
                "b：",
                "Mean",
                "SD",
                "CV",
                "携带污染率CLH",
                "携带污染率CHL",
            ]
        ):
            return False
        if len(normalized) >= 4 and not looks_like_section(first):
            return True
    return False


def _clean_object_title(text: str) -> str:
    value = str(text or "").strip()
    return value.split("\n", 1)[0].strip()


def _extract_object_name(row: list[SourceCell]) -> str:
    texts = compact_texts(cell.text for cell in row)
    if not texts:
        return ""
    joined = " ".join(texts)
    if "试剂名称" in joined:
        values = [text for text in texts if "试剂名称" not in text]
        return values[-1] if values else ""
    return _clean_object_title(texts[0])


def _extract_title(row: list[SourceCell], section: str) -> str:
    texts = compact_texts(cell.text for cell in row)
    if not texts:
        return section
    if "试剂名称" in texts[0] and len(texts) > 1:
        return section
    return _clean_object_title(texts[0])


def _header_norms(rows: list[list[SourceCell]]) -> set[str]:
    result: set[str] = set()
    keywords = [
        "次数",
        "重复次数",
        "测试数据",
        "采光周期",
        "检测项目",
        "要求",
        "结论",
        "吸光度",
        "结果",
        "均值",
        "SD",
        "CV",
        "MAX",
        "MIN",
        "指标",
        "温度值",
        "参数值",
        "校准状态",
    ]
    for row in rows[:5]:
        for cell in row:
            if any(keyword in cell.text for keyword in keywords):
                normalized = normalize_text(cell.text)
                if normalized:
                    result.add(normalized)
    return result


def read_excel_blocks(path: str | Path, module: str = "M1") -> list[ExcelBlock]:
    """Extract blocks by module -> section -> object -> table."""

    workbook_values = load_workbook(str(path), data_only=True)
    workbook_formulas = load_workbook(str(path), data_only=False)
    blocks: list[ExcelBlock] = []

    for worksheet_values in workbook_values.worksheets:
        worksheet_formulas = workbook_formulas[worksheet_values.title]
        max_col = worksheet_values.max_column or 1
        max_row = worksheet_values.max_row or 1
        evaluator = _build_formula_evaluator(worksheet_values, worksheet_formulas)
        compressed = {
            row: _compress_row(worksheet_values, worksheet_formulas, row, max_col, evaluator)
            for row in range(1, max_row + 1)
        }

        current_section = ""
        current_module = module
        section_at_row: dict[int, str] = {}
        module_at_row: dict[int, str] = {}
        anchors: list[int] = []

        for row_number in range(1, max_row + 1):
            row = compressed[row_number]
            if not row:
                section_at_row[row_number] = current_section
                module_at_row[row_number] = current_module
                continue
            module_marker = _module_from_text(_row_text(row)) if _is_module_marker(row) else None
            if module_marker:
                current_module = module_marker
                section_at_row[row_number] = current_section
                module_at_row[row_number] = current_module
                continue
            first = row[0].text
            if looks_like_section(first):
                current_section = canonical_section(first)
                section_at_row[row_number] = current_section
                module_at_row[row_number] = current_module
                continue
            section_at_row[row_number] = current_section
            module_at_row[row_number] = current_module
            if _is_anchor(row):
                anchors.append(row_number)

        section_rows = [
            row
            for row in range(1, max_row + 1)
            if compressed[row] and looks_like_section(compressed[row][0].text)
        ]
        module_rows = [row for row in range(1, max_row + 1) if compressed[row] and _is_module_marker(compressed[row])]

        for index, start in enumerate(anchors):
            candidates = [value for value in anchors[index + 1 :] if value > start]
            candidates.extend(value for value in section_rows if value > start)
            candidates.extend(value for value in module_rows if value > start)
            end = min(candidates) - 1 if candidates else max_row
            while end > start and not compressed[end]:
                end -= 1
            rows = [compressed[row] for row in range(start, end + 1) if compressed[row]]
            if not rows:
                continue
            section = section_at_row.get(start, "")
            blocks.append(
                ExcelBlock(
                    module=module_at_row.get(start, module),
                    sheet=worksheet_values.title,
                    section=section,
                    title=_extract_title(rows[0], section),
                    object_name=_extract_object_name(rows[0]),
                    start_row=start,
                    end_row=end,
                    rows=rows,
                    header_norms=_header_norms(rows),
                )
            )

    unique: list[ExcelBlock] = []
    seen: set[tuple[Any, ...]] = set()
    for block in blocks:
        key = (
            block.sheet,
            block.start_row,
            block.end_row,
            block.module,
            block.section,
            normalize_text(block.object_name),
            normalize_text(block.title),
        )
        if key not in seen:
            seen.add(key)
            unique.append(block)
    return unique
